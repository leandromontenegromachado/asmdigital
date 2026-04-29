import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from sqlalchemy.orm import Session

from app.models import Employee, EvaluationImport, EvaluationImportRow, Review360
from app.services.audit_log_service import AuditLogService


REQUIRED_MAPPING_FIELDS = {
    "evaluated_name",
    "evaluator_email",
    "evaluator_name",
}

SCORE_FIELDS = [
    "general_score",
    "communication_score",
    "teamwork_score",
    "commitment_score",
    "autonomy_score",
    "quality_score",
    "problem_solving_score",
]

COMPETENCY_FIELDS = [field for field in SCORE_FIELDS if field != "general_score"]

ASM_EXTRA_SCORE_HEADERS = [
    "ORIENTACAO PARA RESULTADOS",
    "CONDUTA PROFISSIONAL",
    "CAPACITACAO PARA FUNCAO",
    "CRIATIVIDADE E INOVACAO",
    "PRODUTIVIDADE",
    "CUMPRIMENTO DE PRAZOS METAS",
]

ASM_MANAGER_EVALUATOR_NAMES = {
    "ALESSANDRA MARTINS NUNES",
}

RELATION_ALIASES = {
    "MANAGER": "MANAGER",
    "GESTOR": "MANAGER",
    "GESTOR DIRETO": "MANAGER",
    "PEER": "PEER",
    "PAR": "PEER",
    "COLEGA": "PEER",
    "INTERNAL_CLIENT": "INTERNAL_CLIENT",
    "CLIENTE INTERNO": "INTERNAL_CLIENT",
    "CLIENTE": "INTERNAL_CLIENT",
    "SELF": "SELF",
    "AUTO": "SELF",
    "AUTOAVALIACAO": "SELF",
    "AUTOAVALIAÇÃO": "SELF",
}

TEXT_SCORE_ALIASES = {
    "INSUFICIENTE": 0.0,
    "INSATISFATORIO": 0.0,
    "RUIM": 0.0,
    "REGULAR": 25.0,
    "BOM": 50.0,
    "MUITO BOM": 75.0,
    "MUITO_BOM": 75.0,
    "DESTACADO": 100.0,
    "EXCELENTE": 100.0,
    "OTIMO": 100.0,
    "ÓTIMO": 100.0,
}


@dataclass
class ConfirmImportSummary:
    imported_rows: int
    created_reviews: int


class CsvEvaluationImportService:
    def __init__(self, db: Session, actor_id: int | None = None):
        self.db = db
        self.actor_id = actor_id
        self.audit = AuditLogService(db, actor_id)

    def upload_csv(self, cycle_id: int, file_name: str, content: bytes) -> EvaluationImport:
        headers, rows = self._read_tabular_file(file_name, content)
        if not headers:
            raise ValueError("Arquivo sem cabecalho")

        import_record = EvaluationImport(
            cycle_id=cycle_id,
            file_name=file_name,
            status="UPLOADED",
            uploaded_by=self.actor_id,
            total_rows=0,
        )
        self.db.add(import_record)
        self.db.flush()

        total_rows = 0
        for index, row in enumerate(rows, start=2):
            total_rows += 1
            self.db.add(EvaluationImportRow(
                import_id=import_record.id,
                row_number=index,
                raw_data_json={str(key): value for key, value in row.items() if key is not None},
                status="PENDING",
            ))
        import_record.total_rows = total_rows
        self.audit.register_action("UPLOAD_EVALUATION_CSV", "evaluation_imports", import_record.id, None, {
            "file_name": file_name,
            "total_rows": total_rows,
            "headers": headers,
        })
        return import_record

    def map_columns(self, import_record: EvaluationImport, mapping: dict[str, str]) -> EvaluationImport:
        missing = sorted(field for field in REQUIRED_MAPPING_FIELDS if not mapping.get(field))
        has_score = bool(mapping.get("general_score")) or any(mapping.get(field) for field in COMPETENCY_FIELDS)
        if missing or not has_score:
            details = []
            if missing:
                details.append(f"Campos obrigatorios ausentes: {', '.join(missing)}")
            if not has_score:
                details.append("Mapeie general_score ou ao menos uma nota por competencia")
            raise ValueError("; ".join(details))
        import_record.column_mapping_json = mapping
        import_record.status = "MAPPED"
        self.audit.register_action("MAP_EVALUATION_IMPORT_COLUMNS", "evaluation_imports", import_record.id, None, mapping)
        return import_record

    def validate_import(self, import_record: EvaluationImport) -> EvaluationImport:
        mapping = import_record.column_mapping_json or {}
        if not mapping:
            raise ValueError("Mapeamento de colunas nao informado")

        valid_rows = 0
        invalid_rows = 0
        for row in import_record.rows:
            try:
                row.normalized_data_json = self._normalize_row(row.raw_data_json, mapping)
                row.status = "VALID"
                row.error_message = None
                valid_rows += 1
            except ValueError as exc:
                row.normalized_data_json = None
                row.status = "ERROR"
                row.error_message = str(exc)
                invalid_rows += 1

        import_record.valid_rows = valid_rows
        import_record.invalid_rows = invalid_rows
        import_record.status = "VALIDATED" if invalid_rows == 0 else "ERROR"
        import_record.error_message = None if invalid_rows == 0 else "Existem linhas invalidas"
        self.audit.register_action("VALIDATE_EVALUATION_IMPORT", "evaluation_imports", import_record.id, None, {
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
        })
        return import_record

    def confirm_import(self, import_record: EvaluationImport) -> ConfirmImportSummary:
        if import_record.status not in {"VALIDATED", "ERROR"}:
            raise ValueError("Importacao precisa ser validada antes da confirmacao")
        valid_rows = [row for row in import_record.rows if row.status == "VALID" and row.normalized_data_json]
        created_reviews = 0
        for row in valid_rows:
            data = row.normalized_data_json or {}
            evaluated = self._find_or_create_employee(data["evaluated_email"], data["evaluated_name"])
            evaluator = self._find_or_create_employee(data["evaluator_email"], data["evaluator_name"])
            review = Review360(
                cycle_id=import_record.cycle_id,
                import_id=import_record.id,
                import_row_id=row.id,
                evaluator_id=evaluator.id if evaluator else None,
                evaluator_email=data["evaluator_email"],
                evaluator_name=data["evaluator_name"],
                evaluated_id=evaluated.id,
                evaluated_email=data["evaluated_email"],
                evaluated_name=data["evaluated_name"],
                relation_type=data["relation_type"],
                score=data["general_score"],
                general_score=data["general_score"],
                communication_score=data.get("communication_score"),
                teamwork_score=data.get("teamwork_score"),
                commitment_score=data.get("commitment_score"),
                autonomy_score=data.get("autonomy_score"),
                quality_score=data.get("quality_score"),
                problem_solving_score=data.get("problem_solving_score"),
                strengths_comment=data.get("strengths_comment"),
                improvement_comment=data.get("improvement_comment"),
                general_comment=data.get("general_comment"),
                comment=data.get("general_comment"),
                submitted_at=datetime.fromisoformat(data["submitted_at"]) if data.get("submitted_at") else datetime.utcnow(),
            )
            self.db.add(review)
            row.status = "IMPORTED"
            created_reviews += 1

        import_record.status = "IMPORTED"
        self.audit.register_action("CONFIRM_EVALUATION_IMPORT", "evaluation_imports", import_record.id, None, {
            "imported_rows": len(valid_rows),
            "created_reviews": created_reviews,
        })
        return ConfirmImportSummary(imported_rows=len(valid_rows), created_reviews=created_reviews)

    @staticmethod
    def headers(import_record: EvaluationImport) -> list[str]:
        first_row = import_record.rows[0] if import_record.rows else None
        return list((first_row.raw_data_json or {}).keys()) if first_row else []

    @staticmethod
    def normalize_score(value: Any) -> float | None:
        if value is None or str(value).strip() == "":
            return None
        text = str(value).strip()
        alias = TEXT_SCORE_ALIASES.get(CsvEvaluationImportService._normalize_label(text))
        if alias is not None:
            return alias
        text = text.replace(",", ".")
        try:
            number = float(text)
        except ValueError as exc:
            raise ValueError(f"Nota invalida: {value}") from exc
        if 1 <= number <= 5:
            return round((number - 1) * 25, 2)
        if 0 <= number <= 100:
            return round(number, 2)
        raise ValueError(f"Nota fora da escala permitida: {value}")

    @staticmethod
    def _detect_delimiter(text: str) -> str:
        first_line = text.splitlines()[0] if text.splitlines() else ""
        return ";" if first_line.count(";") > first_line.count(",") else ","

    def _read_tabular_file(self, file_name: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
        extension = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else "csv"
        if extension == "csv":
            text = self._decode_text(content)
            reader = csv.DictReader(StringIO(text), delimiter=self._detect_delimiter(text))
            return list(reader.fieldnames or []), [
                {str(key): value for key, value in row.items() if key is not None}
                for row in reader
            ]
        if extension == "xlsx":
            return self._read_xlsx(content)
        if extension == "xls":
            return self._read_xls(content)
        raise ValueError("Formato nao suportado. Use CSV, XLSX ou XLS.")

    @staticmethod
    def _decode_text(content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError("Nao foi possivel decodificar o arquivo")

    @staticmethod
    def _read_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Suporte a XLSX indisponivel. Instale openpyxl.") from exc
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        return CsvEvaluationImportService._rows_from_matrix(rows)

    @staticmethod
    def _read_xls(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("Suporte a XLS indisponivel. Instale xlrd.") from exc
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        rows = [[sheet.cell_value(rowx, colx) for colx in range(sheet.ncols)] for rowx in range(sheet.nrows)]
        return CsvEvaluationImportService._rows_from_matrix(rows)

    @staticmethod
    def _rows_from_matrix(matrix: list[Any]) -> tuple[list[str], list[dict[str, Any]]]:
        if not matrix:
            return [], []
        headers = [str(value).strip() if value is not None else "" for value in matrix[0]]
        records: list[dict[str, Any]] = []
        for values in matrix[1:]:
            if not values or all(value is None or str(value).strip() == "" for value in values):
                continue
            records.append({
                headers[index]: values[index] if index < len(values) else None
                for index in range(len(headers))
                if headers[index]
            })
        return headers, records

    def _normalize_row(self, raw: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
        def value(field: str) -> str | None:
            column = mapping.get(field)
            raw_value = raw.get(column) if column else None
            return str(raw_value).strip() if raw_value is not None and str(raw_value).strip() else None

        evaluated_name = value("evaluated_name")
        evaluated_email = value("evaluated_email") or self._synthetic_email(evaluated_name)
        evaluator_email = value("evaluator_email")
        evaluator_name = value("evaluator_name")
        relation_type = self._normalize_relation(value("relation_type")) or self._infer_relation(evaluator_name, evaluated_name)

        missing = [
            label for label, current in {
                "evaluated_email": evaluated_email,
                "evaluated_name": evaluated_name,
                "evaluator_email": evaluator_email,
                "evaluator_name": evaluator_name,
                "relation_type": relation_type,
            }.items()
            if not current
        ]
        if missing:
            raise ValueError(f"Campos obrigatorios sem valor: {', '.join(missing)}")

        scores = {field: self.normalize_score(value(field)) for field in SCORE_FIELDS}
        general_score = scores.get("general_score")
        if general_score is None:
            competence_scores = [score for field, score in scores.items() if field != "general_score" and score is not None]
            competence_scores.extend(self._extra_asm_scores(raw, mapping))
            if not competence_scores:
                raise ValueError("Informe general_score ou ao menos uma nota por competencia")
            general_score = round(sum(competence_scores) / len(competence_scores), 2)
        scores["general_score"] = general_score

        submitted_at = self._parse_datetime(value("submitted_at"))
        return {
            "evaluated_email": evaluated_email.lower(),
            "evaluated_name": evaluated_name,
            "evaluator_email": evaluator_email.lower(),
            "evaluator_name": evaluator_name,
            "relation_type": relation_type,
            "submitted_at": submitted_at.isoformat() if submitted_at else None,
            **scores,
            "strengths_comment": value("strengths_comment"),
            "improvement_comment": value("improvement_comment"),
            "general_comment": value("general_comment"),
        }

    @staticmethod
    def _normalize_relation(value: str | None) -> str | None:
        if not value:
            return None
        return RELATION_ALIASES.get(value.strip().upper())

    @classmethod
    def _infer_relation(cls, evaluator_name: str | None, evaluated_name: str | None) -> str:
        if evaluator_name and evaluated_name and cls._normalize_label(evaluator_name) == cls._normalize_label(evaluated_name):
            return "SELF"
        if evaluator_name and cls._normalize_label(evaluator_name) in ASM_MANAGER_EVALUATOR_NAMES:
            return "MANAGER"
        return "PEER"

    @classmethod
    def _extra_asm_scores(cls, raw: dict[str, Any], mapping: dict[str, str]) -> list[float]:
        mapped_columns = {column for column in mapping.values() if column}
        scores: list[float] = []
        for column, raw_value in raw.items():
            if not column or column in mapped_columns:
                continue
            normalized_column = cls._normalize_label(str(column))
            if any(header in normalized_column for header in ASM_EXTRA_SCORE_HEADERS):
                score = cls.normalize_score(raw_value)
                if score is not None:
                    scores.append(score)
        return scores

    @staticmethod
    def _synthetic_email(name: str | None) -> str | None:
        if not name:
            return None
        slug = re.sub(r"[^a-z0-9]+", ".", name.strip().lower()).strip(".")
        return f"{slug or 'avaliado'}@evaluation.asmdigital.com"

    @staticmethod
    def _normalize_label(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().upper())
        return "".join(char for char in normalized if not unicodedata.combining(char))

    @staticmethod
    def _parse_datetime(value: str | None) -> datetime | None:
        if not value:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None

    def _find_or_create_employee(self, email: str | None, name: str | None) -> Employee | None:
        if not email:
            return None
        employee = self.db.query(Employee).filter(Employee.email == email.lower()).first()
        if employee:
            if name and employee.name != name:
                employee.name = name
            return employee
        employee = Employee(name=name or email, email=email.lower(), active=True)
        self.db.add(employee)
        self.db.flush()
        return employee
