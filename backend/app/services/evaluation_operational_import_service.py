import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Employee, EmployeeRhData, EvaluationImport, EvaluationImportRow, PerformanceIndicator
from app.services.evaluation_scoring_service import EvaluationScoringService


EXCLUDED_RPM_AREAS = [
    "DSD ASM ATIVIDADES DE TREINAMENTO",
    "DSD ASM ATIVIDADES GENERICAS",
]

NAME_STOPWORDS = {"A", "DA", "DE", "DO", "DAS", "DOS", "E"}


@dataclass
class OperationalImportSummary:
    imported_rows: int = 0
    updated_employees: int = 0
    updated_indicators: int = 0
    updated_rh_records: int = 0
    warnings: list[str] = field(default_factory=list)
    import_id: int | None = None


class EvaluationOperationalImportService:
    def __init__(self, db: Session, actor_id: int | None = None):
        self.db = db
        self.actor_id = actor_id

    def import_rpm(self, cycle_id: int, file_name: str, content: bytes) -> OperationalImportSummary:
        rows = self._read_xlsx(content)
        summary = OperationalImportSummary(imported_rows=len(rows))
        import_record = self._register_import(cycle_id, file_name, "rpm", rows)
        summary.import_id = import_record.id
        totals: dict[int, dict[str, float]] = {}
        for row in rows:
            name = self._pick(row, ["colaborador", "funcionario", "nome", "recurso", "pessoa"])
            hours = self._number(self._pick(row, ["horas", "hora", "total de horas", "quantidade"]))
            area = self._pick(row, ["projeto", "area", "subarea", "atividade", "classificacao", "item"]) or ""
            if not name or hours is None:
                continue
            employee = self._find_or_create_employee(name)
            summary.updated_employees += 1 if employee.id not in totals else 0
            bucket = totals.setdefault(employee.id, {"total": 0.0, "project": 0.0})
            bucket["total"] += hours
            if not self._is_excluded_rpm_area(area):
                bucket["project"] += hours

        for employee_id, values in totals.items():
            rpm = round((values["project"] / values["total"]) * 100, 2) if values["total"] else None
            self._upsert_indicator(cycle_id, employee_id, rpm=rpm, ihpe=None)
            summary.updated_indicators += 1
        return summary

    def import_ihpe(self, cycle_id: int, file_name: str, content: bytes) -> OperationalImportSummary:
        rows = self._read_xlsx(content)
        summary = OperationalImportSummary(imported_rows=len(rows))
        import_record = self._register_import(cycle_id, file_name, "ihpe", rows)
        summary.import_id = import_record.id
        monthly: dict[int, dict[str, dict[str, float]]] = {}
        current_month: str | None = None
        for row in rows:
            row_month = self._month_key(self._pick(row, ["mes", "competencia", "referencia", "data", "periodo"]))
            if row_month:
                current_month = row_month
            name = self._pick(row, ["colaborador", "funcionario", "nome", "recurso", "pessoa"])
            worked = self._number(self._pick(row, ["horas trabalhadas", "horas pes", "pes", "horas totais", "total de horas"]))
            deliverable = self._number(self._pick(row, [
                "total de horas em entregaveis",
                "horas em entregaveis",
                "horas entregaveis",
                "total de horas em entreg veis",
                "horas em entreg veis",
                "entregaveis",
                "entreg veis",
                "tasks",
                "azure",
            ]))
            if not name or worked is None or deliverable is None:
                continue
            employee = self._find_or_create_employee(name)
            bucket = monthly.setdefault(employee.id, {})
            current = bucket.setdefault(current_month or "periodo", {"worked": 0.0, "deliverable": 0.0})
            current["worked"] += worked
            current["deliverable"] += deliverable

        for employee_id, months in monthly.items():
            percentages = [
                (values["deliverable"] / values["worked"]) * 100
                for values in months.values()
                if values["worked"] > 0
            ]
            ihpe = round(sum(percentages) / len(percentages), 2) if percentages else None
            self._upsert_indicator(cycle_id, employee_id, rpm=None, ihpe=ihpe)
            summary.updated_indicators += 1
        return summary

    def import_rh(self, cycle_id: int, file_name: str, content: bytes) -> OperationalImportSummary:
        rows = self._read_xlsx(content)
        summary = OperationalImportSummary(imported_rows=len(rows))
        import_record = self._register_import(cycle_id, file_name, "rh", rows)
        summary.import_id = import_record.id
        today = date.today()
        for row in rows:
            name = self._pick(row, ["colaborador", "funcionario", "nome", "recurso", "pessoa"])
            if not name:
                continue
            employee = self._find_or_create_employee(name)
            level = self._integer(self._pick(row, ["anc", "nivel", "nivel carreira", "level"]))
            last_merit = self._date(self._pick(row, ["ultimo merito", "ultima promocao", "data ultimo merito", "merito"]))
            admission = self._date(self._pick(row, ["admissao", "data admissao", "data de admissao"]))
            eligible = True
            reason = "Elegivel."
            if level and level >= 24:
                eligible = False
                reason = "Nao elegivel: teto de carreira ANC 24."
            elif last_merit and (today - last_merit).days < 365:
                eligible = False
                reason = "Nao elegivel: promocao por merito ha menos de 12 meses."
            rh = (
                self.db.query(EmployeeRhData)
                .filter(EmployeeRhData.cycle_id == cycle_id, EmployeeRhData.employee_id == employee.id)
                .first()
            )
            if not rh:
                rh = EmployeeRhData(cycle_id=cycle_id, employee_id=employee.id)
                self.db.add(rh)
            rh.career_level = level
            rh.last_merit_date = last_merit
            rh.admission_date = admission
            rh.is_level_one_separate_budget = level == 1
            rh.eligible_for_merit = eligible
            rh.eligibility_reason = reason if level != 1 else "Nivel 1: verba separada; requer desempenho minimo Muito Bom."
            rh.raw_data_json = self._json_safe_row(row)
            summary.updated_rh_records += 1
        return summary

    def _register_import(self, cycle_id: int, file_name: str, kind: str, rows: list[dict[str, Any]]) -> EvaluationImport:
        import_record = EvaluationImport(
            cycle_id=cycle_id,
            file_name=f"[{kind.upper()}] {file_name}",
            status="IMPORTED",
            uploaded_by=self.actor_id,
            column_mapping_json={"type": kind, "source": "operational_import"},
            total_rows=len(rows),
            valid_rows=len(rows),
            invalid_rows=0,
            error_message=None,
        )
        self.db.add(import_record)
        self.db.flush()
        for index, row in enumerate(rows, start=1):
            self.db.add(
                EvaluationImportRow(
                    import_id=import_record.id,
                    row_number=index,
                    raw_data_json=self._json_safe_row(row),
                    normalized_data_json={"type": kind},
                    status="IMPORTED",
                    error_message=None,
                )
            )
        return import_record

    def _upsert_indicator(self, cycle_id: int, employee_id: int, rpm: float | None, ihpe: float | None) -> None:
        indicator = (
            self.db.query(PerformanceIndicator)
            .filter(PerformanceIndicator.cycle_id == cycle_id, PerformanceIndicator.employee_id == employee_id)
            .first()
        )
        if not indicator:
            indicator = PerformanceIndicator(cycle_id=cycle_id, employee_id=employee_id)
            self.db.add(indicator)
        if rpm is not None:
            indicator.rpm_original = rpm
            indicator.rpm_normalized = EvaluationScoringService.normalize_score(rpm)
        if ihpe is not None:
            indicator.ihpe_original = ihpe
            indicator.ihpe_normalized = EvaluationScoringService.normalize_score(ihpe)

    def _find_or_create_employee(self, name: str) -> Employee:
        normalized = self._normalize(name)
        employees = self.db.query(Employee).all()
        for employee in employees:
            if self._normalize(employee.name) == normalized:
                return employee
        for employee in employees:
            if self._is_probable_same_person(employee.name, name):
                return employee
        email = f"{re.sub(r'[^a-z0-9]+', '.', normalized.lower()).strip('.') or 'colaborador'}@evaluation.asmdigital.com"
        employee = Employee(name=name.strip(), email=email, active=True)
        self.db.add(employee)
        self.db.flush()
        return employee

    @classmethod
    def _is_probable_same_person(cls, existing_name: str, imported_name: str) -> bool:
        existing_tokens = cls._name_tokens(existing_name)
        imported_tokens = cls._name_tokens(imported_name)
        if len(existing_tokens) < 2 or len(imported_tokens) < 2:
            return False
        existing_set = set(existing_tokens)
        imported_set = set(imported_tokens)
        shared = existing_set & imported_set
        first_name_matches = existing_tokens[0] == imported_tokens[0]
        last_name_matches = existing_tokens[-1] == imported_tokens[-1]
        if first_name_matches and last_name_matches and len(shared) >= 2:
            return True
        shorter = existing_set if len(existing_set) <= len(imported_set) else imported_set
        return first_name_matches and len(shorter) >= 2 and shorter.issubset(shared)

    @classmethod
    def _name_tokens(cls, value: str) -> list[str]:
        return [token for token in cls._normalize(value).split() if token and token not in NAME_STOPWORDS]

    @staticmethod
    def _read_xlsx(content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        records = []
        for values in rows[1:]:
            if not values or all(value is None or str(value).strip() == "" for value in values):
                continue
            records.append({
                headers[index]: values[index] if index < len(values) else None
                for index in range(len(headers))
                if headers[index]
            })
        return records

    @classmethod
    def _pick(cls, row: dict[str, Any], aliases: list[str]) -> Any:
        normalized_aliases = [cls._normalize(alias) for alias in aliases]
        for key, value in row.items():
            normalized_key = cls._normalize(str(key))
            if any(alias == normalized_key or alias in normalized_key for alias in normalized_aliases):
                return value
        return None

    @staticmethod
    def _normalize(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", str(value).strip().upper())
        normalized = "".join(char for char in normalized if not unicodedata.combining(char))
        return re.sub(r"[^A-Z0-9]+", " ", normalized).strip()

    @classmethod
    def _is_excluded_rpm_area(cls, value: str) -> bool:
        normalized = cls._normalize(value)
        return any(excluded in normalized for excluded in EXCLUDED_RPM_AREAS)

    @staticmethod
    def _number(value: Any) -> float | None:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _integer(value: Any) -> int | None:
        number = EvaluationOperationalImportService._number(value)
        return int(number) if number is not None else None

    @staticmethod
    def _date(value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if not value:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
        return None

    @classmethod
    def _month_key(cls, value: Any) -> str | None:
        parsed = cls._date(value)
        if parsed:
            return parsed.strftime("%Y-%m")
        return str(value).strip() if value else None

    @classmethod
    def _json_safe_row(cls, row: dict[str, Any]) -> dict[str, Any]:
        return {str(key): cls._json_safe_value(value) for key, value in row.items()}

    @classmethod
    def _json_safe_value(cls, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (str, int, float, bool)) or value is None:
            return value
        return str(value)
